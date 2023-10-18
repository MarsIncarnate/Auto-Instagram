import os
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
from openpyxl import load_workbook
from instabot import Bot
import gdown
from PIL import Image
from flask import Flask

app = Flask(__name__)

@app.route('/')
def my_function():
    try:
        # Path to your JSON key file (service account key)
        KEY_FILE = 'insta-automation-398800-fafc4f22d428.json'

        # ID of the Excel file in Google Drive
        FILE_ID = '1UzKTXocecCC6C-xMNqViRB8KH9pSDrv3sP06pUmsfpo'

        # Output file name for downloaded Excel file
        OUTPUT_FILE = 'downloaded_list.xlsx'

        # Specify the range of cells you want to retrieve (e.g., A1 to B5)
        START_CELL = 'B2'
        END_CELL = 'C41'


        # Authenticate using the service account key
        creds = service_account.Credentials.from_service_account_file(KEY_FILE, scopes=['https://www.googleapis.com/auth/drive'])

        # Create a Google Drive API service
        drive_service = build('drive', 'v3', credentials=creds)

        # Download the Excel file
        request = drive_service.files().export_media(fileId=FILE_ID, mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        fh = open(OUTPUT_FILE, 'wb')
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        while not done:
            status, done = downloader.next_chunk()
            print(f"Download {int(status.progress() * 100)}%")

        fh.close()

        # Load the downloaded Excel file using openpyxl
        workbook = load_workbook(OUTPUT_FILE)
        sheet = workbook.active

        # Initialize two separate lists for columns A and B
        column_a_values = []
        column_b_values = []

        # Iterate through the specified range and populate the lists
        for row in sheet.iter_rows(min_row=int(START_CELL[1:]), max_row=int(END_CELL[1:]), values_only=True):
            if row[1] is not None:
                column_a_values.append(row[1])
            if row[2] is not None:
                column_b_values.append(row[2])

        # Print the retrieved values for columns A and B
        print("Values for Column A:")
        print(column_a_values)

        print("\nValues for Column B:")
        print(column_b_values)

        # Retrieve the index from the file
        with open("index.txt", "r") as index_file:
            lpi = int(index_file.read().strip())

        # Process the set of pictures and captions using last_processed_index


        # Google Drive file URL
        google_drive_url = column_a_values[lpi]

        # Destination file path for downloading
        destination_path = 'picture.jpg'

        # Download the image from Google Drive
        gdown.download(google_drive_url, destination_path, quiet=False)

        # Open the downloaded image
        image_path = 'picture.jpg'  
        img = Image.open(image_path)

        # Check the current dimensions
        width, height = img.size

        # Determine the size for the 1:1 aspect ratio
        new_size = min(width, height), min(width, height)

        # Resize the image to the new size
        img = img.resize(new_size)

        # Save the resized image
        output_path = 'pic.jpg'  # Replace with the desired output path
        img.save(output_path)

        # Close the image
        img.close()


        # Now, you can use 'destination_path' to reference the downloaded image in your bot.

        bot = Bot()
        bot.login(username="mars_reloaded", password="profaim", is_threaded=True)

        CAPTION = column_b_values[lpi]

        # Post photos
        # Photos need be resized and, if not in ratio given below.
        # jpg format works more better than others formats.
        # Acceptable Ratio of image:- 90:47, 4:5, 1:1(square image).
        # Keep image and program in same folder.
        # -----------------------------------------------------------
        bot.upload_photo('pic.jpg', caption=str(CAPTION))

        if lpi==39 :
            lpi = 0

        else:
        # Increment the index
            lpi += 1

        # Update the file with the new index
        with open("index.txt", "w") as index_file:
            index_file.write(str(lpi))

        file_path = "pic.jpg.REMOVE_ME"

        if os.path.exists(file_path):
            os.remove(file_path)

        # Your code here
        return "Script executed successfully."
  
    except Exception as e:
        app.logger.error(f"An exception occurred: {str(e)}")
        return "Internal Server Error", 500

if __name__ == '__main__':
    app.run()


